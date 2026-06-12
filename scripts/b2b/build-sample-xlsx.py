# -*- coding: utf-8 -*-
"""B2B営業用サンプルExcel生成: docs/b2b/assets/sample.csv -> sample.xlsx（体裁付き）"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SRC = os.path.join(ROOT, "docs", "b2b", "assets", "sample.csv")
DST = os.path.join(ROOT, "docs", "b2b", "assets", "sample.xlsx")

HEADERS_JA = {
    "pref_id": "都道府県ID", "pref_name": "都道府県", "city_id": "市区町村ID", "city_name": "市区町村",
    "has_subsidy": "制度有無", "subsidy_name": "補助金名称", "max_amount": "上限額", "conditions": "主な交付条件",
    "application_period": "申請期間", "window_name": "担当窓口", "official_url": "自治体公式URL",
    "no_subsidy_note": "備考（制度なし時）", "notes": "備考", "garbage_official_url": "粗大ごみ窓口URL",
}

with open(SRC, encoding="utf-8-sig") as f:
    rows = list(csv.reader(f))

wb = Workbook()
ws = wb.active
ws.title = "解体・空き家補助金サンプル"

# タイトル行
ws.merge_cells("A1:N1")
ws["A1"] = "全国 解体・空き家補助金データベース サンプル（広島県・愛知県）｜株式会社Kogera"
ws["A1"].font = Font(bold=True, size=12)
ws.merge_cells("A2:N2")
ws["A2"] = "全1,726市区町村収録（制度あり844自治体）の一部抜粋。月次更新・出典URL付き。再販・転載不可。"
ws["A2"].font = Font(size=9, color="666666")

header_fill = PatternFill("solid", fgColor="2F5B3F")
header_font = Font(bold=True, color="FFFFFF", size=9)
thin = Border(*[Side(style="thin", color="CCCCCC")]*4)

start = 4
for j, h in enumerate(rows[0], 1):
    c = ws.cell(row=start, column=j, value=HEADERS_JA.get(h, h))
    c.fill, c.font, c.border = header_fill, header_font, thin
    c.alignment = Alignment(horizontal="center", vertical="center")

for i, row in enumerate(rows[1:], start + 1):
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = Font(size=9)
        c.border = thin
        c.alignment = Alignment(vertical="top", wrap_text=(j in (6, 8, 12, 13)))
    if row[4] == "あり":
        ws.cell(row=i, column=5).fill = PatternFill("solid", fgColor="E8F5E9")

widths = [10, 10, 12, 14, 8, 30, 16, 50, 14, 16, 40, 30, 20, 40]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = f"A{start+1}"
ws.auto_filter.ref = f"A{start}:N{start + len(rows) - 1}"

wb.save(DST)
print(f"OK: {DST} ({len(rows)-1} rows)")
